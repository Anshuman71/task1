import sys
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage :python task1.py invoice_filename_without_extension")
    exit(-1)

# get filenames
inPut = sys.argv[1]
outPut = 'SoftDataUpload'

# intialise input file
invoice = load_workbook(inPut + '.xlsx')
invoice_sheet = invoice.active

# initialise output file
sdu = load_workbook(outPut + '.xlsx')
sdu_sheet = sdu.active

# find where to start
startPoint = 1
for i in range(1, 100):
    if sdu_sheet.cell(i, 1).value == None:
        startPoint = i
        break


# initialise helpers data
COD = load_workbook('COD.xlsx')
COD_sheet = COD.active

PPD = load_workbook('PPD.xlsx')
PPD_sheet = PPD.active

GST = load_workbook('GST.xlsx')
GST_sheet = GST.active

# details for first invoice from invoice file

name = invoice_sheet.cell(14, 3).value
sdu_sheet.cell(startPoint, 4).value = name

address = invoice_sheet.cell(15, 3).value
sdu_sheet.cell(startPoint, 5).value = address
city = address.split(', ')
sdu_sheet.cell(startPoint, 6).value = city[3]

phone = invoice_sheet.cell(16, 3).value
phone = phone[9:]
sdu_sheet.cell(startPoint, 9).value = phone

order_no = invoice_sheet.cell(11, 2).value
sdu_sheet.cell(startPoint, 2).value = order_no

invoice_no = invoice_sheet.cell(8, 2).value
sdu_sheet.cell(startPoint, 14).value = invoice_no

gst_seller = invoice_sheet.cell(5, 2).value
gst_code = gst_seller.split('- ')
sdu_sheet.cell(startPoint, 16).value = gst_code[1]

date = invoice_sheet.cell(8, 9).value
sdu_sheet.cell(startPoint, 15).value = date

mode = invoice_sheet.cell(8, 4).value
mode = mode.split(' - ')
mode = mode[0]

sdu_sheet.cell(startPoint, 3).value = mode

awb_code = None

if mode == 'COD':
    for i in range(2, 300):
        if COD_sheet.cell(i, 2).value != 'Used':
            awb_code = COD_sheet.cell(i, 1).value
            COD_sheet.cell(i, 2).value = 'Used'
            COD.save('COD.xlsx')
            break
else:
    for i in range(20, 250):
        if PPD_sheet.cell(i, 2).value != 'Used':
            awb_code = PPD_sheet.cell(i, 1).value
            PPD_sheet.cell(i, 2).value = 'Used'
            PPD.save('PPD.xlsx')
            break

sdu_sheet.cell(startPoint, 1).value = awb_code

description = invoice_sheet.cell(23, 2).value
sdu_sheet.cell(startPoint, 10).value = description

quantity = invoice_sheet.cell(23, 7).value
sdu_sheet.cell(startPoint, 11).value = quantity

rate = invoice_sheet.cell(23, 8).value
sdu_sheet.cell(startPoint, 18).value = rate

tax = invoice_sheet.cell(28, 10).value
sdu_sheet.cell(startPoint, 19).value = tax

state = None

state_code = invoice_sheet.cell(18, 4).value
for i in range(2, 39):
    if GST_sheet.cell(i, 3).value == state_code:
        state = GST_sheet.cell(i, 2).value
sdu_sheet.cell(startPoint, 8).value = state


sdu_sheet.cell(startPoint, 17).value = 'HR IGST'

declared_value = invoice_sheet.cell(29, 10).value
sdu_sheet.cell(startPoint, 13).value = declared_value

if mode == "PPD":
    payable_amount = 0
else:
    payable_amount = declared_value
sdu_sheet.cell(startPoint, 12).value = payable_amount


# details for next invoice from the invoice file


name = invoice_sheet.cell(14, 15).value
sdu_sheet.cell(startPoint + 1, 4).value = name

address = invoice_sheet.cell(15, 15).value
sdu_sheet.cell(startPoint + 1, 5).value = address
city = address.split(', ')
sdu_sheet.cell(startPoint + 1, 6).value = city[3]

phone = invoice_sheet.cell(16, 15).value
phone = phone[9:]
sdu_sheet.cell(startPoint + 1, 9).value = phone

invoice_no = invoice_sheet.cell(8, 14).value
sdu_sheet.cell(startPoint + 1, 14).value = invoice_no


order_no = invoice_sheet.cell(11, 14).value
sdu_sheet.cell(startPoint + 1, 2).value = order_no

gst_seller = invoice_sheet.cell(5, 14).value
gst_code = gst_seller.split('- ')
sdu_sheet.cell(startPoint + 1, 16).value = gst_code[1]

date = invoice_sheet.cell(8, 21).value
sdu_sheet.cell(startPoint + 1, 15).value = date


mode = invoice_sheet.cell(8, 16).value
mode = mode.split(' - ')
mode = mode[0]

sdu_sheet.cell(startPoint + 1, 3).value = mode


if mode == 'COD':
    for i in range(2, 300):
        if COD_sheet.cell(i, 2).value != 'Used':
            awb_code = COD_sheet.cell(i, 1).value
            COD_sheet.cell(i, 2).value = 'Used'
            COD.save('COD.xlsx')
            COD.close()
            break
else:
    for i in range(20, 251):
        if PPD_sheet.cell(i, 2).value != 'Used':
            awb_code = PPD_sheet.cell(i, 1).value
            PPD_sheet.cell(i, 2).value = 'Used'
            PPD.save('PPD.xlsx')
            PPD.close()
            break

sdu_sheet.cell(startPoint + 1, 1).value = awb_code


description = invoice_sheet.cell(23, 14).value
sdu_sheet.cell(startPoint + 1, 10).value = description

quantity = invoice_sheet.cell(23, 19).value
sdu_sheet.cell(startPoint + 1, 11).value = quantity

rate = invoice_sheet.cell(23, 20).value
sdu_sheet.cell(startPoint + 1, 18).value = rate * quantity

tax = invoice_sheet.cell(28, 22).value
sdu_sheet.cell(startPoint + 1, 19).value = tax

sdu_sheet.cell(startPoint + 1, 17).value = 'HR IGST'

declared_value = invoice_sheet.cell(29, 22).value
sdu_sheet.cell(startPoint + 1, 13).value = declared_value

if mode == "PPD":
    payable_amount = 0
else:
    payable_amount = declared_value
sdu_sheet.cell(startPoint + 1, 12).value = payable_amount

state_code = invoice_sheet.cell(18, 16).value
for i in range(2, 39):
    if GST_sheet.cell(i, 3).value == state_code:
        state = GST_sheet.cell(i, 2).value

sdu_sheet.cell(startPoint + 1, 8).value = state

# save file as xlsx
sdu.save(outPut + '.xlsx')
